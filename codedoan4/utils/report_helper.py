# import os
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import PatternFill
# from datetime import datetime
# import re

# class ExcelReport:
#     def __init__(self, file_path, sheet_name="Sheet1", 
#                  tieu_de_base="Báo cáo", header=None):
#         self.file_path = file_path
#         self.sheet_name = sheet_name
#         self.tieu_de_base = tieu_de_base
#         self.header = header  # header tuỳ chỉnh cho mỗi chức năng

#         if os.path.exists(file_path):
#             self.wb = load_workbook(file_path)
#             if sheet_name in self.wb.sheetnames:
#                 self.ws = self.wb[sheet_name]
#             else:
#                 self.ws = self.wb.create_sheet(sheet_name)
#         else:
#             self.wb = Workbook()
#             self.ws = self.wb.active
#             self.ws.title = sheet_name

#         self.tieu_de = self._get_next_tieu_de()
        
# #tìm và thêm tiêu đề mới
#     def _get_next_tieu_de(self):
#         count = 0
#         pattern = re.compile(rf"{re.escape(self.tieu_de_base)} (\d+)")
#         for row in self.ws.iter_rows(min_row=1, max_col=1):
#             cell_val = str(row[0].value)
#             match = pattern.match(cell_val)
#             if match:
#                 count = max(count, int(match.group(1)))

#         new_number = count + 1
#         self.ws.append([f"{self.tieu_de_base} {new_number}"])
#         self._write_header()
#         return f"{self.tieu_de_base} {new_number}"
    
# #ghi dòng tiêu đề cột.
#     def _write_header(self):
#         # Nếu header không được truyền thì dùng mặc định
#         header = self.header or [
#             "STT", "Thời gian", "Thông tin", "Mật khẩu",
#             "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái"
#         ]
#         self.ws.append(header)
        
# #thêm dữ liệu từng dòng, tự chèn thời gian và tô màu trạng thái.
#     def add_row(self, *values):
#         now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#         row = list(values)
#         # nếu cột 2 là thời gian thì tự chèn thời gian
#         if len(row) > 1:
#             row.insert(1, now)
#         self.ws.append(row)
#         # nếu có cột "Trạng thái" thì tô màu
#         status_col = len(row)
#         status = row[-1]
#         fill = PatternFill(
#             start_color="C6EFCE" if status.lower() == "pass" else "FFC7CE",
#             end_color="C6EFCE" if status.lower() == "pass" else "FFC7CE",
#             fill_type="solid",
#         )
#         self.ws[f"{chr(64 + status_col)}{self.ws.max_row}"].fill = fill

#     def save(self):
#         os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
#         self.wb.save(self.file_path)

import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from zipfile import BadZipFile


class ExcelReport:
    def __init__(self, file_path, sheet_name="Sheet1",
                 tieu_de_base="Báo cáo", header=None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.tieu_de_base = tieu_de_base
        self.header = header

        # ===== LOAD FILE – NẾU HỎNG THÌ TẠO LẠI =====
        try:
            if os.path.exists(file_path):
                self.wb = load_workbook(file_path)
            else:
                raise FileNotFoundError
        except (KeyError, BadZipFile, FileNotFoundError):
            # 👉 File Excel bị hỏng → xoá → tạo mới
            if os.path.exists(file_path):
                os.remove(file_path)

            self.wb = Workbook()
            ws = self.wb.active
            ws.title = sheet_name
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            self.wb.save(file_path)

        # ===== LOAD / TẠO SHEET =====
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.create_sheet(sheet_name)

        # ===== THÊM TIÊU ĐỀ =====
        self.tieu_de = self._get_next_tieu_de()
        self.save()

    # ===============================
    # Tạo tiêu đề báo cáo mới
    # ===============================
    def _get_next_tieu_de(self):
        count = 0
        pattern = re.compile(rf"{re.escape(self.tieu_de_base)} (\d+)")

        for row in self.ws.iter_rows(min_row=1, max_col=1):
            if row[0].value:
                match = pattern.match(str(row[0].value))
                if match:
                    count = max(count, int(match.group(1)))

        new_number = count + 1
        self.ws.append([f"{self.tieu_de_base} {new_number}"])
        self._write_header()
        return f"{self.tieu_de_base} {new_number}"

    # ===============================
    # Ghi header
    # ===============================
    def _write_header(self):
        header = self.header or [
            "STT", "Thời gian", "Thông tin",
            "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái"
        ]
        self.ws.append(header)

    # ===============================
    # Ghi 1 dòng dữ liệu
    # ===============================
    def add_row(self, *values):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = list(values)

        if len(row) > 1:
            row.insert(1, now)

        self.ws.append(row)

        # tô màu trạng thái
        status = str(row[-1]).lower()
        col_letter = get_column_letter(len(row))
        cell = self.ws[f"{col_letter}{self.ws.max_row}"]

        fill = PatternFill(
            start_color="C6EFCE" if status == "pass" else "FFC7CE",
            end_color="C6EFCE" if status == "pass" else "FFC7CE",
            fill_type="solid"
        )
        cell.fill = fill

    # ===============================
    # Lưu file
    # ===============================
    def save(self):
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        self.wb.save(self.file_path)
